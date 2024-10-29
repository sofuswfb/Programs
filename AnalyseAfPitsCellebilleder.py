# -*- coding: utf-8 -*-
'''
The script only works on images taken at 4x zoom on the microscope in Tandlaegeskolen
This script will use the multifluor and the UV images taken on this microscope

This needs to be done manually:
    Renaming files to make them compatible with Nicjolas' script (no spaces)
    Execution of Nicjolas' script to generate images from all 3 channels
    Execution of the rotation script to generate images where the substrate is at a right angle

'''

# Start of script. With good rpecision this script can hit sub 10 pixel precision

import numpy as np
import cv2
import math
import openpyxl

# Dictionary which contain the orientation of masters used.
# An item consist of the key, which is given as a master name (like M5 for amster 5)
# and a value, which will be the orientation of the master. See attatched images in
# the script folder and match these to the master using a microscope.
# Possibilities for oritentations are: N, W, S and E.
master_dictionary = {'W': 'W', 'M7':'W'}

# List of corners for calibration
corners = []
# Variable for x and y values one is currently looking at on images
current_x, current_y = 0, 0
# Minimum zoom level to prevent distortion
min_zoom_level = 10

# Functions
def save_image_coordinates(event, x, y, flags, param):
    """
    Mouse callback function to get coordinates relative to the full image.
    Can only collect one (x,y) entry
    """
    if event == cv2.EVENT_LBUTTONDOWN:
        if len(corners) < 3:
            # Calculate the coordinates relative to the full image
            zoom_factor = zoom_level / 100.0
            full_x = int(current_x + x / zoom_factor)
            full_y = int(current_y + y / zoom_factor)
            corners.append((full_x, full_y))
            print(f"Saved coordinates ({full_x}, {full_y})")

def save_image_coordinates_infinite_entries(event, x, y, flags, param):
    """
    Mouse callback function to get coordinates relative to the full image.
    Can collect and infinite amount of (x,y) entries
    """
    if event == cv2.EVENT_LBUTTONDOWN:
        # Calculate the coordinates relative to the full image
        zoom_factor = zoom_level / 100.0
        full_x = int(current_x + x / zoom_factor)
        full_y = int(current_y + y / zoom_factor)
        gemte_felt_koordinater.append((full_x, full_y))
        print(f"Saved coordinates ({full_x}, {full_y})")

def treatment_of_saved_fields(liste):
    '''
    Takes a list of coordiantes and matches them to a field.
    Returns a list of the fields which one wants to get data from in the excel file.
    Only relevant if one chooses to choose which fields one wants to keep
    '''
    feltnumre = []
    for coords in liste:
        for i in range(1,65):
            x_low = cell_count_matching.get(i)[1][0]
            x_high = cell_count_matching.get(i)[1][0] + square_size
            y_low = cell_count_matching.get(i)[1][1]
            y_high = cell_count_matching.get(i)[1][1] + square_size
            
            if x_low < coords[0] < x_high and y_low < coords[1] < y_high and i not in feltnumre:
                feltnumre.append(i)
                break
    return feltnumre

def update_image(x, y, zoom, image, window_size, window_name):
    """
    Update the displayed image based on the trackbar positions and zoom level.
    """
    global current_x, current_y, zoom_level
    current_x, current_y = x, y
    zoom_level = max(zoom, min_zoom_level)  # Ensure zoom level is not below the minimum
    
    h, w = image.shape[:2]
    
    # Calculate the size of the displayed area based on zoom level
    zoom_factor = zoom / 100.0
    display_w = int(window_size / zoom_factor)
    display_h = int(window_size / zoom_factor)

    # Ensure the displayed area does not exceed the image dimensions
    x_end = min(x + display_w, w)
    y_end = min(y + display_h, h)
    cropped_image = image[y:y_end, x:x_end]

    # Resize the cropped image to fit the window size
    resized_image = cv2.resize(cropped_image, (window_size, window_size))
    cv2.imshow(window_name, resized_image)

def on_trackbar(val, image, window_size, window_name):
    """
    Callback function for trackbar.
    """
    x = cv2.getTrackbarPos('X', window_name)
    y = cv2.getTrackbarPos('Y', window_name)
    zoom = cv2.getTrackbarPos('Zoom', window_name)
    
    # Ensure the zoom level does not go below the minimum zoom level
    if zoom < min_zoom_level:
        zoom = min_zoom_level
        cv2.setTrackbarPos('Zoom', window_name, min_zoom_level)
    
    update_image(x, y, zoom, image, window_size, window_name)
    
def change_orientation(orientation, cell_counts):
    '''
    Takes the list of fields and the cell counts corresponding to the fields,
    and changes the number of the fields to properly match dependent on the
    orientation of the master used.
    '''
    output = {}
    i = 1

    if orientation == 'N':
        for coords in cell_counts:
            output[i] = (cell_counts.get(coords), coords)
            i += 1
        return output

    if orientation == 'W':
        for coords in cell_counts:
            output[i] = (cell_counts.get(coords), coords)
            if i < 57:
                i += 8
            else:
                i -= 55
        return output
    
    if orientation == 'E':
        i = 64
        for coords in cell_counts:
            output[i] = (cell_counts.get(coords), coords)
            if i < 9:
                i += 55
            else:
                i -= 8
        return output

    if orientation == 'S':
        i = 57
        for coords in cell_counts:
            output[i] = (cell_counts.get(coords), coords)
            if i//8 == i/8:
                i -= 15
            else:
                i += 1
        return output



#%% Choice of fields for calibration

# 'master' and 'sample_folder' should be the only variables you change between different samples
master = 'W'
sample_folder = 'Your cell data/Imprints(PCmedTI)/dag_5/_Dag_5_sample_2_/'
# This path shoudl be chosen so the image 'Overlay Cellebilleder nummereret.png' is found here
# along with sample folders for all your cell experiments
path = 'C:/Users/User/Desktop/Andreas script - Data folder/Andreas script - Data folder/'

# Load the base image
base_image_path = 'Zbehandlet - Rotated BF image.tif'
base_image = cv2.imread(path + sample_folder + base_image_path)

# Load the mask image
mask_image_path = 'Overlay Cellebilleder nummereret.png'
mask_image = cv2.imread(path + mask_image_path)

# Checking that both images are found
assert base_image is not None, "Error loading base_image"
assert mask_image is not None, "Error loading mask_image"

print('Press the top left corner of a field and remember the field number. '+
      'Choose a field close to the middle for best precision. '+
      'Close the image when you have clicked a corner.')

# Resize the mask image to match the base image dimensions (if necessary)
mask_image = cv2.resize(mask_image, (base_image.shape[1], base_image.shape[0]))

# Blend the images
alpha = 0.5  # Transparency factor for the mask image
blended_image = cv2.addWeighted(base_image, 1 - alpha, mask_image, alpha, 0)

# Window size is chosen, edit this if the images are displayed on a small screen
window_size = 800

# Picture is shown
cv2.imshow('Blended Image', blended_image[:window_size, :window_size])
cv2.setMouseCallback('Blended Image', save_image_coordinates)

# Trackbars are created making it possible to change X, Y and Zoom.
cv2.createTrackbar('X', 'Blended Image', 0, max(0, blended_image.shape[1] - window_size), lambda val: on_trackbar(val, blended_image, window_size, 'Blended Image'))
cv2.createTrackbar('Y', 'Blended Image', 0, max(0, blended_image.shape[0] - window_size), lambda val: on_trackbar(val, blended_image, window_size, 'Blended Image'))
cv2.createTrackbar('Zoom', 'Blended Image', 100, 500, lambda val: on_trackbar(val, blended_image, window_size, 'Blended Image'))  # Zoom range from 100% to 500%
cv2.waitKey(0)
cv2.destroyAllWindows()

valgt_felt = int(input('What is the field number of the chosen field: '))

print('Now choose the top left corner of a field close to the edge of the sample. '+
      'Again remember the field number. Close the image when you have clicked a corner.')

# Picture is shown
cv2.imshow('Blended Image', blended_image[:window_size, :window_size])
cv2.setMouseCallback('Blended Image', save_image_coordinates)
cv2.createTrackbar('X', 'Blended Image', 0, max(0, blended_image.shape[1] - window_size), lambda val: on_trackbar(val, blended_image, window_size, 'Blended Image'))
cv2.createTrackbar('Y', 'Blended Image', 0, max(0, blended_image.shape[0] - window_size), lambda val: on_trackbar(val, blended_image, window_size, 'Blended Image'))
cv2.createTrackbar('Zoom', 'Blended Image', 100, 500, lambda val: on_trackbar(val, blended_image, window_size, 'Blended Image'))  # Zoom range from 100% to 500%
cv2.waitKey(0)
cv2.destroyAllWindows()

kali_1 = int(input('What is the field number of the chosen field: '))

print('Now choose a field as far away from the most recently chosen field for best results. '+
      'Again click the top left corner and remember the field number. Close the image when you have clicked a corner.')

# Picture is shown
cv2.imshow('Blended Image', blended_image[:window_size, :window_size])
cv2.setMouseCallback('Blended Image', save_image_coordinates)
cv2.createTrackbar('X', 'Blended Image', 0, max(0, blended_image.shape[1] - window_size), lambda val: on_trackbar(val, blended_image, window_size, 'Blended Image'))
cv2.createTrackbar('Y', 'Blended Image', 0, max(0, blended_image.shape[0] - window_size), lambda val: on_trackbar(val, blended_image, window_size, 'Blended Image'))
cv2.createTrackbar('Zoom', 'Blended Image', 100, 500, lambda val: on_trackbar(val, blended_image, window_size, 'Blended Image'))  # Zoom range from 100% to 500%
cv2.waitKey(0)
cv2.destroyAllWindows()

kali_2 = int(input('What is the field number of the chosen field: '))

#%% Calculation of field positions (Calibration)

assert len(corners) == 3, '3 corners were not selected, please select 3 corners in the first part of the script'
assert type(valgt_felt) is int, 'Matching field numbers were not found, remember to write the field number after selecting a field'

# Manuel conversion ratios
# pixels_to_mm = 1/389
# mm_to_pixels = 389

# Calculation of the x and y length between the corners chosen for calibration
# The space between fields is set to 2.5 mm.
cols_apart = abs((kali_1 - 8*(math.floor(kali_1/8))) - (kali_2 - 8*(math.floor(kali_2/8))))
felt_plus_inter_pixels_x = round(abs(corners[1][0] - corners[2][0]) / cols_apart)
rows_apart = abs(math.ceil(kali_1/8) - math.ceil(kali_2/8))
felt_plus_inter_pixels_y = round(abs(corners[1][1] - corners[2][1]) / rows_apart)

# Finding the top left corner of field 1, followed by collection of all top left corners
felt_1_corner = corners[0]
# y-coordinate for field 1
while valgt_felt - 8 > 0:
    valgt_felt -= 8
    felt_1_corner = (felt_1_corner[0], felt_1_corner[1] - felt_plus_inter_pixels_y)
# x-coordinate for field 1
while valgt_felt - 1 > 0:
    valgt_felt -= 1
    felt_1_corner = (felt_1_corner[0] - felt_plus_inter_pixels_x, felt_1_corner[1])

# Collection of all top left corners
felt = 0
all_corners = []
# Loop through rows
for row in range(8):
    # Loop through columns
    for col in range(8):
        felt += 1

        all_corners.append((felt_1_corner[0] + felt_plus_inter_pixels_x * col, felt_1_corner[1] + felt_plus_inter_pixels_y * row))

# Checking for negative values, if none are found a confirmation of the corners being 
# found is given
for i in all_corners:
    if i[0] < 0 or i[1] < 1:
        
        assert False, 'Negative coordinates detected, try again with better calibration'
    elif i == all_corners[63] and len(all_corners) == 64:
        print('The Coordinates of the top left corenrs for all 64 fields are found')



#%% Cell counting

# Define the size of each field
inter_space_pixels = 157
square_size = round((felt_plus_inter_pixels_x + felt_plus_inter_pixels_y) / 2) - inter_space_pixels


# Load the image
UV_image_path = 'Zbehandlet - Rotated UV image.tif'
UV_image = cv2.imread(path + sample_folder + UV_image_path)

# Check if the image was loaded successfully
assert UV_image is not None, 'Error: UV image was not succesfully loaded'

# Define HSV range for cell detection
hsv_lower = np.array([0, 0, 0])
hsv_upper = np.array([31, 31, 31])

def count_cells_in_square_hsv(square, hsv_lower, hsv_upper):
    """
    Count the number of cells in the given square using HSV color space.
    
    Parameters:
    square (numpy.ndarray): BGR image of the square region.
    hsv_lower (numpy.ndarray): Lower bound of HSV values for cell detection.
    hsv_upper (numpy.ndarray): Upper bound of HSV values for cell detection.
    
    Returns:
    int: Number of cells detected in the square.
    """
    # Convert the square to HSV color space
    hsv = cv2.cvtColor(square, cv2.COLOR_BGR2HSV)

    # Create a binary mask using the given HSV range
    mask = cv2.inRange(hsv, hsv_lower, hsv_upper)

    # Invert the binary mask to ensure cells are white and background is black
    inverted_mask = cv2.bitwise_not(mask)

    # Apply morphological operations to clean up the mask
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    opening = cv2.morphologyEx(inverted_mask, cv2.MORPH_OPEN, kernel, iterations=1)
    closing = cv2.morphologyEx(opening, cv2.MORPH_CLOSE, kernel, iterations=2)

    # Find contours in the binary mask
    contours, _ = cv2.findContours(closing, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # Parameters for cell area FOR REVIEW MAKE THESE FOR EACH DAY DUE TO CHANGE IN CELL SIZE
    minimum_area = 10
    average_cell_area = 60
    connected_cell_area = 150
    too_large_area = 200
    # Count the number of cells based on contour areas
    cells = 0
    for contour in contours:
        area = cv2.contourArea(contour)
        if area > minimum_area:
            if area > too_large_area:
                continue
            elif area > connected_cell_area:
                cells += math.floor(area / average_cell_area)
            else:
                cells += 1

    return cells

# Dictionary to store the cell count for each square
cell_counts = {}

# Process each square
for i, (x, y) in enumerate(all_corners):
    # Extract the square from the image, y og x er benbart sdan her
    square = UV_image[y:y+square_size, x:x+square_size]

    # Count the cells in the square using HSV color space
    cell_count = count_cells_in_square_hsv(square, hsv_lower, hsv_upper)

    # Store the cell count in the dictionary
    cell_counts[(x, y)] = cell_count


# Remove # if you want to print all coordinates and their cell counts
# for coords, count in cell_counts.items():
#     print(f'Square at {coords}: {count} cells')

# Remove # if you want the full UV image with squares drawn on the image
for (x, y) in all_corners:
    cv2.rectangle(UV_image, (x, y), (x + square_size, y + square_size), (255, 0, 0), 2)
cv2.imshow('Image with Squares', UV_image[:window_size, :window_size])
cv2.createTrackbar('X', 'Image with Squares', 0, max(0, UV_image.shape[1] - window_size), lambda val: on_trackbar(val, UV_image, window_size, 'Image with Squares'))
cv2.createTrackbar('Y', 'Image with Squares', 0, max(0, UV_image.shape[0] - window_size), lambda val: on_trackbar(val, UV_image, window_size, 'Image with Squares'))
cv2.createTrackbar('Zoom', 'Image with Squares', 100, 500, lambda val: on_trackbar(val, UV_image, window_size, 'Image with Squares'))  # Zoom range from 100% to 500%
cv2.waitKey(0)
cv2.destroyAllWindows()


#%% Data extraction

orientering = master_dictionary.get(master)

# The orientation is used to change the number of the fields to match the proper orientation
cell_count_matching = change_orientation(orientering, cell_counts)

# Path to the existing Excel file and the path to save the new file
existing_file_path = path + 'CellCountDataSkabelon.xlsx'
new_file_path = path + sample_folder + 'CellCountData.xlsx'

# Load the existing Excel file
wb = openpyxl.load_workbook(existing_file_path)

# Select the active sheet (or specify the sheet name if you know it)
ws = wb.active

# Inserts the name of the sample in the top of the file
ws['A1'] = sample_folder.split('/')[-2]

Y_for_valg_N_for_alle = 0
while Y_for_valg_N_for_alle not in ['y', 'n', 'Y', 'N']:
    Y_for_valg_N_for_alle = str(input("Type 'Y' if you want to manually choose which fields are extracted"+
                                      ", type 'N' to keep all fields: "))

if Y_for_valg_N_for_alle in ['y', 'Y']:
    # Image is opened and you can manually click which fields you want to extract
    gemte_felt_koordinater = []
    print('Choose the fields you want to save in the Excel file')
    cv2.imshow('Base Image', base_image[:window_size, :window_size])
    cv2.setMouseCallback('Base Image', save_image_coordinates_infinite_entries)
    cv2.createTrackbar('X', 'Base Image', 0, max(0, base_image.shape[1] - window_size), lambda val: on_trackbar(val, base_image, window_size, 'Base Image'))
    cv2.createTrackbar('Y', 'Base Image', 0, max(0, base_image.shape[0] - window_size), lambda val: on_trackbar(val, base_image, window_size, 'Base Image'))
    cv2.createTrackbar('Zoom', 'Base Image', 100, 500, lambda val: on_trackbar(val, base_image, window_size, 'Base Image'))  # Zoom range from 100% to 500%
    cv2.waitKey(0)
    cv2.destroyAllWindows()

    # Calculation of the fields numbers one wants to keep based on the coordinates chosen through clicking
    felter_der_beholdes = treatment_of_saved_fields(gemte_felt_koordinater)
    for i in [1, 8, 57, 64]:
        if i not in felter_der_beholdes:
            felter_der_beholdes.append(i)
    print('The following fields were chosen: ' + str(sorted(felter_der_beholdes)))

else:
    # If one chooses to keep all fields
    felter_der_beholdes = range(1,65)

# Write cell counts to cells B2 to B65
for i in felter_der_beholdes:
    # B2 to B65
    cell_address = f'B{i+1}'
    # Get cell count for field number (or 0 if not found)
    ws[cell_address] = cell_count_matching.get(i, 0)[0]

# Save the edited Excel file to the new location
try:
    wb.save(new_file_path)
    print()
    print(f"Excel file saved to {new_file_path}")
except PermissionError:
    print(f"PermissionError: Make sure the file {new_file_path} before trying to edit it.")


#%% For debugging, can display a single square. Remember coordinates in cv2 are (y, x)
# square = UV_image[750:750+square_size, 686:686+square_size]

def debug_single_square():
    kopi = square
    original = kopi.copy()
    hsv = cv2.cvtColor(kopi, cv2.COLOR_BGR2HSV)
    
    hsv_lower = np.array([0,0,0])
    hsv_upper = np.array([31,31,31])
    mask = cv2.inRange(hsv, hsv_lower, hsv_upper)
    inverted_mask = cv2.bitwise_not(mask)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3,3))
    opening = cv2.morphologyEx(inverted_mask, cv2.MORPH_OPEN, kernel, iterations=1)
    close = cv2.morphologyEx(opening, cv2.MORPH_CLOSE, kernel, iterations=2)
    
    cnts = cv2.findContours(close, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0] if len(cnts) == 2 else cnts[1]
    
    minimum_area = 10
    average_cell_area = 60
    connected_cell_area = 150
    too_large_area = 200
    cells = 0
    for c in cnts:
        area = cv2.contourArea(c)
        if area > minimum_area:
            cv2.drawContours(original, [c], -1, (36,255,12), 2)
            if area > too_large_area:
                continue
            elif area > connected_cell_area:
                cells += math.floor(area / average_cell_area)
            else:
                cells += 1
    print('Cells: {}'.format(cells))
    cv2.imshow('close', close)
    cv2.imshow('original', original)
    cv2.waitKey(0)

# If you want to run the debug function remove the # under this line
# debug_single_square()

